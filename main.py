"""
Bot de gastos personales con Google Sheets — multi usuario, listo para producción
Requiere: pip install python-telegram-bot gspread google-auth python-dotenv
"""

import os
import io
import json
import logging
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from telegram import Update
from telegram.ext import (
    ApplicationBuilder, CommandHandler, ContextTypes,
    ConversationHandler, MessageHandler, filters,
    PicklePersistence
)
from google.oauth2.service_account import Credentials
import gspread
from dotenv import load_dotenv

load_dotenv()

# ─── Logging ──────────────────────────────────────────────────────────────────
logging.basicConfig(
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s",
    level=logging.INFO
)
logger = logging.getLogger(__name__)

# ─── Google Sheets ─────────────────────────────────────────────────────────────
SCOPES = ['https://www.googleapis.com/auth/spreadsheets']

def iniciar_sheets():
    """Inicializa la conexión a Google Sheets desde variable de entorno o archivo."""
    creds_json = os.environ.get("GOOGLE_CREDENTIALS")
    if creds_json:
        info = json.loads(creds_json)
        creds = Credentials.from_service_account_info(info, scopes=SCOPES)
    else:
        creds = Credentials.from_service_account_file('credentials.json', scopes=SCOPES)
    return gspread.authorize(creds)

client      = iniciar_sheets()
SHEET_ID    = os.environ.get("SHEET_ID")
spreadsheet = client.open_by_key(SHEET_ID)
ENCABEZADO  = ["Fecha", "Monto", "Categoría", "Descripción"]

# ─── Rate limiting ─────────────────────────────────────────────────────────────
RATE_LIMIT: dict[int, datetime] = {}
RATE_LIMIT_SEGUNDOS = 2

def esta_en_rate_limit(chat_id: int) -> bool:
    ahora = datetime.now()
    ultima = RATE_LIMIT.get(chat_id)
    if ultima and (ahora - ultima).total_seconds() < RATE_LIMIT_SEGUNDOS:
        return True
    RATE_LIMIT[chat_id] = ahora
    return False

# ─── Gestión de pestañas por usuario ──────────────────────────────────────────
def obtener_hoja_usuario(chat_id: int):
    nombre = str(chat_id)
    try:
        return spreadsheet.worksheet(nombre)
    except gspread.exceptions.WorksheetNotFound:
        hoja = spreadsheet.add_worksheet(title=nombre, rows=1000, cols=10)
        hoja.append_row(ENCABEZADO)
        return hoja

# ─── Estados del ConversationHandler ──────────────────────────────────────────
MONTO, CATEGORIA, DESCRIPCION = range(3)

# ─── Helpers ──────────────────────────────────────────────────────────────────
def fila_a_dict(fila: list) -> dict | None:
    try:
        return {
            "fecha":       fila[0],
            "monto":       float(fila[1]),
            "categoria":   fila[2],
            "descripcion": fila[3] if len(fila) > 3 else "—",
        }
    except (IndexError, ValueError):
        return None

def obtener_gastos(chat_id: int) -> list[dict]:
    try:
        hoja  = obtener_hoja_usuario(chat_id)
        filas = hoja.get_all_values()
        if len(filas) <= 1:
            return []
        return [g for g in (fila_a_dict(f) for f in filas[1:]) if g]
    except Exception as e:
        logger.error(f"Error obteniendo gastos para {chat_id}: {e}")
        return None

def mes_actual() -> str:
    return datetime.now().strftime("%m/%Y")

def formatear_monto(monto: float) -> str:
    return f"CAD ${monto:,.2f}"

# ─── /start ───────────────────────────────────────────────────────────────────
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if esta_en_rate_limit(update.effective_chat.id):
        return

    nombre = update.effective_user.first_name
    try:
        obtener_hoja_usuario(update.effective_chat.id)
    except Exception as e:
        logger.error(f"Error creando hoja para {update.effective_chat.id}: {e}")
        await update.message.reply_text("⚠️ Error al inicializar tu cuenta. Intenta de nuevo.")
        return

    await update.message.reply_text(
        f"👋 Hola, *{nombre}*. Bienvenido a tu bot de gastos personales.\n\n"
        "🔒 Tus datos son privados y solo tú puedes verlos.\n\n"
        "📋 *Comandos disponibles:*\n\n"
        "/gasto — Registrar un nuevo gasto\n"
        "/resumen — Ver total gastado\n"
        "/mes — Ver gastos del mes actual\n"
        "/categoria — Ver gastos por categoría\n"
        "/descargar — Descargar tus gastos en Excel\n"
        "/ayuda — Mostrar este mensaje",
        parse_mode="Markdown"
    )

# ─── /gasto ───────────────────────────────────────────────────────────────────
async def gasto_inicio(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if esta_en_rate_limit(update.effective_chat.id):
        return ConversationHandler.END
    await update.message.reply_text("💵 ¿Cuánto gastaste? (solo el número, ej: 45.50)\n\nEscribe /cancelar para salir.")
    return MONTO

async def gasto_monto(update: Update, context: ContextTypes.DEFAULT_TYPE):
    texto = update.message.text.replace(",", ".").replace("$", "").strip()
    try:
        monto = float(texto)
        if monto <= 0 or monto > 1_000_000:
            raise ValueError
    except ValueError:
        await update.message.reply_text("⚠️ Ingresa un número válido entre 0 y 1,000,000. Ej: 45.50")
        return MONTO

    context.user_data["monto"] = monto
    await update.message.reply_text("🏷️ ¿Cuál es la categoría? (ej: Comida, Transporte, Salud)")
    return CATEGORIA

async def gasto_categoria(update: Update, context: ContextTypes.DEFAULT_TYPE):
    categoria = update.message.text.strip().capitalize()[:50]  # max 50 chars
    context.user_data["categoria"] = categoria
    await update.message.reply_text("📝 Agrega una descripción breve (o /saltar para omitir)")
    return DESCRIPCION

async def gasto_descripcion(update: Update, context: ContextTypes.DEFAULT_TYPE):
    descripcion = update.message.text.strip()[:200]  # max 200 chars
    return await guardar_gasto(update, context, descripcion)

async def gasto_sin_descripcion(update: Update, context: ContextTypes.DEFAULT_TYPE):
    return await guardar_gasto(update, context, "—")

async def guardar_gasto(update: Update, context: ContextTypes.DEFAULT_TYPE, descripcion: str):
    chat_id   = update.effective_chat.id
    monto     = context.user_data.get("monto")
    categoria = context.user_data.get("categoria")
    fecha     = datetime.now().strftime("%d/%m/%Y")

    if not monto or not categoria:
        await update.message.reply_text("⚠️ Error en el registro. Intenta de nuevo con /gasto.")
        context.user_data.clear()
        return ConversationHandler.END

    try:
        hoja = obtener_hoja_usuario(chat_id)
        hoja.append_row([fecha, monto, categoria, descripcion])
    except Exception as e:
        logger.error(f"Error guardando gasto para {chat_id}: {e}")
        await update.message.reply_text("⚠️ Error al guardar en Google Sheets. Intenta de nuevo.")
        context.user_data.clear()
        return ConversationHandler.END

    await update.message.reply_text(
        f"✅ *Gasto registrado*\n\n"
        f"📅 Fecha: {fecha}\n"
        f"💵 Monto: {formatear_monto(monto)}\n"
        f"🏷️ Categoría: {categoria}\n"
        f"📝 Descripción: {descripcion}",
        parse_mode="Markdown"
    )
    context.user_data.clear()
    return ConversationHandler.END

async def cancelar(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data.clear()
    await update.message.reply_text("❌ Registro cancelado.")
    return ConversationHandler.END

# ─── /resumen ─────────────────────────────────────────────────────────────────
async def resumen(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if esta_en_rate_limit(update.effective_chat.id):
        return

    gastos = obtener_gastos(update.effective_chat.id)

    if gastos is None:
        await update.message.reply_text("⚠️ Error al conectar con Google Sheets. Intenta más tarde.")
        return
    if not gastos:
        await update.message.reply_text("📭 No tienes gastos registrados aún.")
        return

    total = sum(g["monto"] for g in gastos)
    por_categoria = {}
    for g in gastos:
        cat = g["categoria"]
        por_categoria[cat] = por_categoria.get(cat, 0) + g["monto"]

    cats = "\n".join(
        f"  • {cat}: {formatear_monto(m)}"
        for cat, m in sorted(por_categoria.items(), key=lambda x: -x[1])
    )

    await update.message.reply_text(
        f"📊 *Tu resumen total de gastos*\n\n"
        f"💰 *Total gastado:* {formatear_monto(total)}\n"
        f"🔢 *N° de gastos:* {len(gastos)}\n\n"
        f"*Por categoría:*\n{cats}",
        parse_mode="Markdown"
    )

# ─── /mes ─────────────────────────────────────────────────────────────────────
async def mes(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if esta_en_rate_limit(update.effective_chat.id):
        return

    gastos = obtener_gastos(update.effective_chat.id)

    if gastos is None:
        await update.message.reply_text("⚠️ Error al conectar con Google Sheets. Intenta más tarde.")
        return

    mes_act    = mes_actual()
    gastos_mes = [g for g in gastos if g["fecha"].endswith(mes_act)]

    if not gastos_mes:
        await update.message.reply_text(f"📭 No tienes gastos registrados en {mes_act}.")
        return

    total  = sum(g["monto"] for g in gastos_mes)
    lineas = "\n".join(
        f"  {i+1}. [{g['categoria']}] {formatear_monto(g['monto'])} — {g['descripcion']} ({g['fecha']})"
        for i, g in enumerate(gastos_mes)
    )

    await update.message.reply_text(
        f"📅 *Tus gastos de {mes_act}*\n\n"
        f"{lineas}\n\n"
        f"💰 *Total del mes:* {formatear_monto(total)}",
        parse_mode="Markdown"
    )

# ─── /categoria ───────────────────────────────────────────────────────────────
async def categoria(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if esta_en_rate_limit(update.effective_chat.id):
        return

    gastos = obtener_gastos(update.effective_chat.id)

    if gastos is None:
        await update.message.reply_text("⚠️ Error al conectar con Google Sheets. Intenta más tarde.")
        return
    if not gastos:
        await update.message.reply_text("📭 No tienes gastos registrados aún.")
        return

    por_categoria = {}
    for g in gastos:
        cat = g["categoria"]
        if cat not in por_categoria:
            por_categoria[cat] = {"total": 0, "count": 0}
        por_categoria[cat]["total"] += g["monto"]
        por_categoria[cat]["count"] += 1

    total_general = sum(g["monto"] for g in gastos)
    lineas = []
    for cat, datos in sorted(por_categoria.items(), key=lambda x: -x[1]["total"]):
        porcentaje = (datos["total"] / total_general) * 100
        lineas.append(
            f"*{cat}*\n"
            f"  💵 {formatear_monto(datos['total'])} ({porcentaje:.1f}%)\n"
            f"  🔢 {datos['count']} gasto(s)"
        )

    await update.message.reply_text(
        "🏷️ *Tus gastos por categoría*\n\n" + "\n\n".join(lineas),
        parse_mode="Markdown"
    )


# ─── /descargar ───────────────────────────────────────────────────────────────
async def descargar(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if esta_en_rate_limit(update.effective_chat.id):
        return

    chat_id = update.effective_chat.id
    gastos  = obtener_gastos(chat_id)

    if gastos is None:
        await update.message.reply_text("⚠️ Error al conectar con Google Sheets. Intenta más tarde.")
        return
    if not gastos:
        await update.message.reply_text("📭 No tienes gastos registrados para descargar.")
        return

    await update.message.reply_text("⏳ Generando tu archivo Excel...")

    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Mis Gastos"

        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        header_alig = Alignment(horizontal="center", vertical="center")

        headers    = ["Fecha", "Monto (CAD)", "Categoría", "Descripción"]
        col_widths = [14, 14, 20, 40]

        for col, (header, width) in enumerate(zip(headers, col_widths), start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alig
            ws.column_dimensions[cell.column_letter].width = width

        ws.row_dimensions[1].height = 22

        for i, g in enumerate(gastos, start=2):
            ws.cell(row=i, column=1, value=g["fecha"])
            ws.cell(row=i, column=2, value=round(g["monto"], 2))
            ws.cell(row=i, column=3, value=g["categoria"])
            ws.cell(row=i, column=4, value=g["descripcion"])
            if i % 2 == 0:
                fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
                for c in range(1, 5):
                    ws.cell(row=i, column=c).fill = fill

        fila_total = len(gastos) + 2
        ws.cell(row=fila_total, column=1, value="TOTAL").font = Font(bold=True)
        ws.cell(row=fila_total, column=2, value=round(sum(g["monto"] for g in gastos), 2)).font = Font(bold=True)

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        nombre_archivo = f"gastos_{datetime.now().strftime('%Y%m')}.xlsx"
        await update.message.reply_document(
            document=buffer,
            filename=nombre_archivo,
            caption=(
                f"📊 Tus gastos — {len(gastos)} registro(s)\n"
                f"💰 Total: CAD ${sum(g['monto'] for g in gastos):,.2f}"
            )
        )

    except Exception as e:
        logger.error(f"Error generando Excel para {chat_id}: {e}")
        await update.message.reply_text("⚠️ Error al generar el archivo. Intenta de nuevo.")

# ─── /ayuda ───────────────────────────────────────────────────────────────────
async def ayuda(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(
        "🤖 *Comandos disponibles:*\n\n"
        "/gasto — Registrar un nuevo gasto\n"
        "/resumen — Ver tu total gastado y desglose\n"
        "/mes — Ver tus gastos del mes actual\n"
        "/categoria — Ver tus gastos por categoría\n"
        "/descargar — Descargar tus gastos en Excel\n"
        "/cancelar — Cancelar el registro en curso\n"
        "/ayuda — Mostrar este mensaje",
        parse_mode="Markdown"
    )

# ─── Error handler global ─────────────────────────────────────────────────────
async def error_handler(update: object, context: ContextTypes.DEFAULT_TYPE):
    logger.error(f"Error no controlado: {context.error}", exc_info=context.error)
    if isinstance(update, Update) and update.effective_message:
        await update.effective_message.reply_text(
            "⚠️ Ocurrió un error inesperado. Por favor intenta de nuevo."
        )

# ─── Main ─────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    TOKEN = os.environ.get("TELEGRAM_TOKEN")

    # Persistencia en disco: sobrevive reinicios del servidor
    persistence = PicklePersistence(filepath="bot_data.pkl")

    app = ApplicationBuilder().token(TOKEN).persistence(persistence).build()

    conv = ConversationHandler(
        entry_points=[CommandHandler("gasto", gasto_inicio)],
        states={
            MONTO: [MessageHandler(filters.TEXT & ~filters.COMMAND, gasto_monto)],
            CATEGORIA: [MessageHandler(filters.TEXT & ~filters.COMMAND, gasto_categoria)],
            DESCRIPCION: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, gasto_descripcion),
                CommandHandler("saltar", gasto_sin_descripcion),
            ],
        },
        fallbacks=[CommandHandler("cancelar", cancelar)],
        persistent=True,
        name="gasto_conv",
    )

    app.add_handler(conv)
    app.add_handler(CommandHandler("start",     start))
    app.add_handler(CommandHandler("resumen",   resumen))
    app.add_handler(CommandHandler("mes",       mes))
    app.add_handler(CommandHandler("categoria", categoria))
    app.add_handler(CommandHandler("descargar", descargar))
    app.add_handler(CommandHandler("ayuda",     ayuda))
    app.add_error_handler(error_handler)

    logger.info("💰 Bot de gastos iniciado...")
    app.run_polling(allowed_updates=Update.ALL_TYPES)
